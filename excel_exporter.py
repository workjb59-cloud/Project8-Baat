"""
Excel Exporter for Boutiqaat products
Creates organized Excel files with proper structure
"""
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List
import io
import logging
import config

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export scraped data to Excel files"""
    
    def __init__(self):
        """Initialize the exporter"""
        self.columns = config.PRODUCT_COLUMNS
    
    def create_styled_workbook(self, df: pd.DataFrame, sheet_name: str = 'Products') -> Workbook:
        """Create a styled Excel workbook from a DataFrame"""
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with reasonable limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        return wb
    
    def prepare_dataframe(self, products: List[Dict]) -> pd.DataFrame:
        """Prepare product data as a DataFrame with proper column order"""
        if not products:
            return pd.DataFrame(columns=self.columns)
        
        df = pd.DataFrame(products)
        
        # Ensure all expected columns exist
        for col in self.columns:
            if col not in df.columns:
                df[col] = None
        
        # Reorder columns
        available_columns = [col for col in self.columns if col in df.columns]
        df = df[available_columns]
        
        return df
    
    def export_celebrities_to_excel(self, celebrities_data: Dict[str, List[Dict]]) -> Dict[str, bytes]:
        """
        Export celebrities data - one Excel file per celebrity
        
        Args:
            celebrities_data: Dict mapping celebrity names to product lists
            
        Returns:
            Dict mapping filenames to file bytes
        """
        excel_files = {}
        
        for celebrity_name, products in celebrities_data.items():
            if not products:
                logger.warning(f"No products found for celebrity: {celebrity_name}")
                continue
            
            logger.info(f"Creating Excel for celebrity: {celebrity_name} ({len(products)} products)")
            
            # Prepare DataFrame
            df = self.prepare_dataframe(products)
            
            # Create styled workbook
            wb = self.create_styled_workbook(df, sheet_name=celebrity_name[:31])  # Excel sheet name limit
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Clean filename
            filename = f"{self._clean_filename(celebrity_name)}.xlsx"
            excel_files[filename] = excel_buffer.getvalue()
        
        return excel_files
    
    def export_brands_to_excel(self, brands_data: Dict[str, List[Dict]]) -> Dict[str, bytes]:
        """
        Export brands data - one Excel file with all brands as sheets
        
        Args:
            brands_data: Dict mapping brand names to product lists
            
        Returns:
            Dict with single filename mapping to file bytes
        """
        logger.info(f"Creating single Excel file for {len(brands_data)} brands")
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Track if any sheets were added
        sheets_added = 0
        
        for brand_name, products in brands_data.items():
            if not products:
                logger.warning(f"No products found for brand: {brand_name}")
                continue
            
            logger.info(f"Adding sheet for brand: {brand_name} ({len(products)} products)")
            
            # Prepare DataFrame
            df = self.prepare_dataframe(products)
            
            # Create sheet name (Excel limit is 31 characters)
            sheet_name = self._clean_sheet_name(brand_name)[:31]
            ws = wb.create_sheet(title=sheet_name)
            sheets_added += 1
            
            # Write data with styling
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Style header row
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF", size=11)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                    
                    # Add borders
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        # Only save if we have sheets
        if sheets_added == 0:
            logger.warning("No products found for any brand, skipping Excel file")
            return {}
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return {'all_brands.xlsx': excel_buffer.getvalue()}
    
    def export_categories_to_excel(self, categories_data: Dict[str, List[Dict]]) -> Dict[str, bytes]:
        """
        Export categories data - one Excel file per main category, subcategories as sheets
        
        Args:
            categories_data: Dict mapping category names to product lists
            
        Returns:
            Dict mapping filenames to file bytes
        """
        excel_files = {}
        
        # Group categories by main category
        grouped_categories = self._group_categories(categories_data)
        
        for main_category, subcategories in grouped_categories.items():
            logger.info(f"Creating Excel for category: {main_category} ({len(subcategories)} subcategories)")
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Track if any sheets were added
            sheets_added = 0
            
            for subcat_name, products in subcategories.items():
                if not products:
                    logger.warning(f"No products found for subcategory: {subcat_name}")
                    continue
                
                logger.info(f"Adding sheet for subcategory: {subcat_name} ({len(products)} products)")
                
                # Prepare DataFrame
                df = self.prepare_dataframe(products)
                
                # Create sheet name
                sheet_name = self._clean_sheet_name(subcat_name)[:31]
                ws = wb.create_sheet(title=sheet_name)
                sheets_added += 1
                
                # Write data with styling
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        if r_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF", size=11)
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        else:
                            cell.alignment = Alignment(vertical="top", wrap_text=True)
                        
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.border = thin_border
                
                # Auto-adjust columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                ws.freeze_panes = 'A2'
            
            # Only save if we have sheets
            if sheets_added == 0:
                logger.warning(f"No products found for any subcategory in {main_category}, skipping Excel file")
                continue
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            filename = f"{self._clean_filename(main_category)}.xlsx"
            excel_files[filename] = excel_buffer.getvalue()
        
        return excel_files
    
    def _group_categories(self, categories_data: Dict[str, List[Dict]]) -> Dict[str, Dict[str, List[Dict]]]:
        """Group categories by main category"""
        grouped = {}
        
        for category_name, products in categories_data.items():
            # Extract main category (first part before '/')
            parts = category_name.split('/')
            main_category = parts[0].strip() if parts else category_name
            
            if main_category not in grouped:
                grouped[main_category] = {}
            
            grouped[main_category][category_name] = products
        
        return grouped
    
    def _clean_filename(self, name: str) -> str:
        """Clean filename for saving"""
        # Remove invalid characters
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Limit length
        return name[:100]
    
    def _clean_sheet_name(self, name: str) -> str:
        """Clean sheet name for Excel"""
        # Excel doesn't allow these characters in sheet names
        invalid_chars = '[]:*?/\\'
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        return name


if __name__ == '__main__':
    # Test the exporter
    exporter = ExcelExporter()
    
    # Sample data
    sample_products = [
        {
            'entity_id': '123',
            'sku': 'MU-001',
            'name': 'Test Product',
            'brand_name': 'Test Brand',
            'regular_price': '10.000',
            'final_price': '8.000',
            'currency_code': 'KWD'
        }
    ]
    
    celebrities_data = {'Test Celebrity': sample_products}
    files = exporter.export_celebrities_to_excel(celebrities_data)
    
    print(f"Generated {len(files)} Excel files")
